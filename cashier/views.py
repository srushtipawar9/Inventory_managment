from django.shortcuts import render, redirect, reverse, get_object_or_404
from decimal import Decimal
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.forms import formset_factory, modelformset_factory
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Q
import json
from . models import (
    DaftarBarang,
    DaftarTransaksi,
    ListProductTransaksi,
    EstimateSlip,
    EstimateSlipItem,
    FinancialInsightReport,
)
from accounts.models import Profile
from cashier.forms import (
    DaftarBarangForm,
    TransaksiProductListForm,
    ListProductTransaksiForm,
    DaftarTransaksiForm,
)
from django.utils.timezone import datetime
from data.models import JCBPart, Vendor


# ── Helper: is the current user an admin? ────────────────────────────────────
def is_admin(user):
    return user.is_staff or user.is_superuser or user.username == 'admin'


def handler404(request):
    return render(request, '404.html', status=404)

def handler500(request):
    return render(request, '500.html', status=500)


@login_required()
def HomeIndex(request):
    today = datetime.today()

    # Revenue – today's sales total
    txns_today = DaftarTransaksi.objects.filter(
        created__day=today.day,
        created__month=today.month,
        created__year=today.year,
        user_id=request.user.profile.id,
    )
    pendapatan_hari_ini = sum(Decimal(t.total) for t in txns_today)

    # Inventory summary
    inventory_qs = DaftarBarang.objects.filter(user_id=request.user.profile.id, status='Active')
    total_items = inventory_qs.count()
    total_stock_value = sum(Decimal(d.subtotal_harga_beli or 0) for d in inventory_qs)
    total_sell_value  = sum(Decimal(d.subtotal_harga_jual or 0) for d in inventory_qs)
    potential_profit  = total_sell_value - total_stock_value

    # All-time transactions
    all_txns = DaftarTransaksi.objects.filter(user_id=request.user.profile.id)
    total_sales = sum(Decimal(t.total) for t in all_txns)

    context = {
        'data_pendapatan': pendapatan_hari_ini,
        'total_items': total_items,
        'total_stock_value': total_stock_value,
        'total_sell_value': total_sell_value,
        'potential_profit': potential_profit,
        'total_sales': total_sales,
        'txns_today_count': txns_today.count(),
        'recent_items': inventory_qs.order_by('-created')[:10],
        'is_admin': is_admin(request.user),
    }
    return render(request, 'cashier/home.html', context)


def _part_choices(stocks, prefill=''):
    choices = [('', '-- Select Item --')]
    for stock in stocks:
        label = f"{stock.part_number} - {stock.name}"
        choices.append((label, label))
    return choices


@login_required()
def InputStock(request):
    DaftarBarangFormset = formset_factory(DaftarBarangForm, extra=1)
    stocks = JCBPart.objects.all().order_by('category', 'name')
    prefill = request.GET.get('prefill', '')
    part_choices = _part_choices(stocks, prefill)
    vendor_names = list(
        Vendor.objects.order_by('city', 'name').values_list('name', flat=True).distinct()
    )
    company_list = ['JCB', 'CAT', 'Komatsu', 'Volvo', 'Tata Hitachi', 'CASE', 'Doosan', 'Liebherr']

    if request.method == 'POST':
        post_data = request.POST.copy()

        total_forms = int(post_data.get('form-TOTAL_FORMS', 0))
        filled_form_indices = []
        for i in range(total_forms):
            nama  = post_data.get(f'form-{i}-nama_product', '').strip()
            qty   = post_data.get(f'form-{i}-jumlah_produk', '').strip()
            price = post_data.get(f'form-{i}-harga_beli_satuan', '').strip()
            if nama or qty or price:
                filled_form_indices.append(i)

        clean_post_data = post_data.copy()
        clean_post_data['form-TOTAL_FORMS']   = str(len(filled_form_indices))
        clean_post_data['form-INITIAL_FORMS'] = '0'
        clean_post_data['form-MIN_NUM_FORMS'] = '0'
        clean_post_data['form-MAX_NUM_FORMS'] = '1000'

        for key in list(clean_post_data.keys()):
            if key.startswith('form-') and not any(
                key.startswith(f'form-{prefix}')
                for prefix in ['TOTAL_FORMS', 'INITIAL_FORMS', 'MIN_NUM_FORMS', 'MAX_NUM_FORMS']
            ):
                del clean_post_data[key]

        clean_files = request.FILES.copy()
        for key in list(request.FILES.keys()):
            if key.startswith('form-'):
                del clean_files[key]

        for new_idx, old_idx in enumerate(filled_form_indices):
            clean_post_data[f'form-{new_idx}-user'] = str(request.user.profile.id)
            for field_name in [
                'nama_product', 'part_for_what', 'hsn_sac', 'jumlah_produk', 'vendor',
                'company', 'harga_beli_satuan', 'gst_percent', 'gst_amount', 'amt_incl_tax',
                'laba_persen', 'harga_jual_satuan', 'mrp', 'status',
            ]:
                old_key = f'form-{old_idx}-{field_name}'
                new_key = f'form-{new_idx}-{field_name}'
                if old_key in post_data:
                    val = post_data[old_key]
                    if field_name in ['harga_beli_satuan', 'harga_jual_satuan', 'mrp']:
                        val = val.replace(',', '').strip()
                    clean_post_data[new_key] = val

            file_key = f'form-{old_idx}-image'
            if file_key in request.FILES:
                clean_files[f'form-{new_idx}-image'] = request.FILES[file_key]

        formset_post = DaftarBarangFormset(
            clean_post_data,
            clean_files,
            form_kwargs={'part_choices': part_choices},
        )
        if formset_post.is_valid():
            saved = 0
            for form in formset_post:
                if not form.cleaned_data.get('nama_product'):
                    continue
                jumlah = form.cleaned_data.get('jumlah_produk', 0) or 0
                harga_beli_raw = form.data.get(form.add_prefix('harga_beli_satuan'), '0')
                try:
                    harga = Decimal(str(harga_beli_raw).replace(',', '').strip())
                except Exception:
                    harga = Decimal('0')
                if jumlah < 1 or harga < 1:
                    messages.warning(
                        request,
                        f'Qty and Purchase Price required for {form.cleaned_data.get("nama_product")}',
                    )
                    continue
                form.save()
                saved += 1
            if saved:
                messages.success(request, f'Inventory saved successfully! ({saved} item(s)) saved to permanent stock.')
                return redirect('TotalStock')
            messages.warning(request, 'No valid rows to save. Please fill Item Name, Qty and Purchase Price.')
        else:
            messages.warning(request, 'Please correct the highlighted fields and try again.')

        context = {
            'stocks': stocks,
            'forms': formset_post,
            'request_user': request.user.profile.id,
            'prefill': prefill,
            'vendor_names': vendor_names,
            'company_list': company_list,
        }
        return render(request, 'cashier/input_data.html', context)

    formset = DaftarBarangFormset(
        initial=[{'user': request.user.profile.id}],
        form_kwargs={'part_choices': part_choices}
    )
    context = {
        'stocks': stocks,
        'forms': formset,
        'request_user': request.user.profile.id,
        'prefill': prefill,
        'vendor_names': vendor_names,
        'company_list': company_list,
    }
    return render(request, 'cashier/input_data.html', context)


@login_required()
def TotalStock(request):
    if is_admin(request.user):
        data = DaftarBarang.objects.filter(user_id=request.user.profile.id)
    else:
        data = DaftarBarang.objects.filter(user_id=request.user.profile.id, status='Active')
    context = {
        'data': data,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'cashier/stock.html', context)


@login_required()
def EditStock(request, pk):
    item = get_object_or_404(DaftarBarang, nomor=pk, user_id=request.user.profile.id)
    stocks = JCBPart.objects.all().order_by('category', 'name')
    part_choices = _part_choices(stocks)
    vendor_names = list(
        Vendor.objects.order_by('city', 'name').values_list('name', flat=True).distinct()
    )
    company_list = ['JCB', 'CAT', 'Komatsu', 'Volvo', 'Tata Hitachi', 'CASE', 'Doosan', 'Liebherr']

    if request.method == 'POST':
        post_data = request.POST.copy()
        for key in ['harga_beli_satuan', 'harga_jual_satuan', 'mrp']:
            if key in post_data:
                post_data[key] = post_data[key].replace(',', '').strip()
        form = DaftarBarangForm(post_data, request.FILES, instance=item, part_choices=part_choices)
        if form.is_valid():
            # Only admin can change status
            if not is_admin(request.user):
                form.instance.status = item.status
            form.save()
            messages.success(request, f"Inventory item '{item.nama_product}' updated successfully!")
            return redirect('TotalStock')
        else:
            messages.warning(request, "Please correct the errors in the form.")
    else:
        form = DaftarBarangForm(instance=item, part_choices=part_choices)

    context = {
        'form': form,
        'item': item,
        'stocks': stocks,
        'vendor_names': vendor_names,
        'company_list': company_list,
        'is_admin': is_admin(request.user),
    }
    return render(request, 'cashier/edit_stock.html', context)


@login_required()
def DeleteStock(request, pk):
    item = get_object_or_404(DaftarBarang, nomor=pk, user_id=request.user.profile.id)
    name = item.nama_product
    item.delete()
    messages.success(request, f"Inventory item '{name}' deleted successfully!")
    return redirect('TotalStock')


@login_required()
def BulkUploadStock(request):
    """
    Bulk upload inventory stock via Excel / CSV.
    Expected columns (order does not matter, matched by header name):
        Item Name | No. / For What | HSN / SAC | Qty | Vendor | Company |
        Purchase  | GST %          | Profit %  | MRP
    Auto-calculated: GST Amt, Incl. Tax, Sales Price
    """
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.warning(request, 'Please select a file to upload.')
            return redirect('BulkUploadStock')

        ext = uploaded_file.name.rsplit('.', 1)[-1].lower()
        try:
            if ext == 'csv':
                import csv, io
                decoded = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            elif ext in ('xls', 'xlsx'):
                import openpyxl
                from decimal import Decimal as D
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                messages.warning(request, 'Unsupported format. Please upload .csv, .xls or .xlsx')
                return redirect('BulkUploadStock')
        except Exception as e:
            messages.error(request, f'Error reading file: {e}')
            return redirect('BulkUploadStock')

        # Normalize header → field mapping (flexible, case-insensitive)
        FIELD_MAP = {
            'item name':      'nama_product',
            'name':           'nama_product',
            'no. / for what': 'part_for_what',
            'no/for what':    'part_for_what',
            'no./for what':   'part_for_what',
            'for what':       'part_for_what',
            'hsn / sac':      'hsn_sac',
            'hsn/sac':        'hsn_sac',
            'hsn':            'hsn_sac',
            'qty':            'jumlah_produk',
            'quantity':       'jumlah_produk',
            'vendor':         'vendor',
            'company':        'company',
            'purchase':       'harga_beli_satuan',
            'purchase price': 'harga_beli_satuan',
            'gst %':          'gst_percent',
            'gst%':           'gst_percent',
            'gst':            'gst_percent',
            'profit %':       'laba_persen',
            'profit%':        'laba_persen',
            'profit':         'laba_persen',
            'mrp':            'mrp',
        }

        saved = 0
        errors = []
        from decimal import Decimal, InvalidOperation

        for i, row in enumerate(rows, start=2):  # row 2 onwards (row 1 = header)
            # Normalize keys
            normalized = {}
            for k, v in row.items():
                key = str(k).strip().lower() if k else ''
                field = FIELD_MAP.get(key)
                if field:
                    normalized[field] = str(v).strip() if v is not None else ''

            name = normalized.get('nama_product', '').strip()
            if not name:
                continue  # skip empty rows

            def to_decimal(val, default='0'):
                try:
                    return Decimal(str(val).replace(',', '').strip() or default)
                except InvalidOperation:
                    return Decimal(default)

            def to_int(val, default=0):
                try:
                    return int(str(val).replace(',', '').strip() or default)
                except (ValueError, TypeError):
                    return default

            qty      = to_int(normalized.get('jumlah_produk', 1), 1)
            purchase = to_decimal(normalized.get('harga_beli_satuan', '0'))
            gst_pct  = to_decimal(normalized.get('gst_percent', '0'))
            profit   = to_int(normalized.get('laba_persen', 10), 10)
            mrp      = to_decimal(normalized.get('mrp', '0'))

            if qty < 1 or purchase <= 0:
                errors.append(f'Row {i} ({name}): Qty and Purchase Price are required.')
                continue

            try:
                DaftarBarang.objects.create(
                    user=request.user.profile,
                    nama_product=name,
                    part_for_what=normalized.get('part_for_what', ''),
                    hsn_sac=normalized.get('hsn_sac', ''),
                    jumlah_produk=qty,
                    vendor=normalized.get('vendor', ''),
                    company=normalized.get('company', ''),
                    harga_beli_satuan=purchase,
                    gst_percent=gst_pct,
                    laba_persen=profit,
                    mrp=mrp,
                    status='Active',
                )
                saved += 1
            except Exception as e:
                errors.append(f'Row {i} ({name}): {e}')

        if saved:
            messages.success(request, f'✅ Bulk upload successful! {saved} item(s) added to inventory.')
        if errors:
            for err in errors[:5]:
                messages.warning(request, err)
        if not saved and not errors:
            messages.warning(request, 'No valid data rows found. Check your file format.')

        return redirect('TotalStock')

    return render(request, 'cashier/bulk_upload_stock.html')


@login_required()
def DownloadSampleInventoryCSV(request):
    """Return a sample CSV file with the exact column headers for bulk stock upload."""
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_inventory_upload.csv"'
    writer = csv.writer(response)
    # Header row — exact names the bulk upload parser recognises
    writer.writerow([
        'Item Name', 'No. / For What', 'HSN / SAC',
        'Qty', 'Vendor', 'Company',
        'Purchase', 'GST %', 'Profit %', 'MRP',
    ])
    # Two demo rows
    writer.writerow([
        'Hydraulic Hose Assembly', 'jcb-x100 / Boom Cylinder', '84314995',
        '2', 'Nagpur Earthmovers', 'JCB',
        '10000', '18', '10', '13000',
    ])
    writer.writerow([
        'Bucket Pin Kit', 'Dipper Arm', '84313990',
        '5', 'Pune Spares', 'CAT',
        '2500', '18', '12', '3200',
    ])
    return response



@login_required()
def HideStock(request, pk):
    if not is_admin(request.user):
        messages.error(request, "You do not have permission to hide inventory records.")
        return redirect('TotalStock')
    item = get_object_or_404(DaftarBarang, nomor=pk)
    item.status = 'Hidden'
    item.save()
    messages.success(request, f"'{item.nama_product}' has been hidden.")
    return redirect('TotalStock')


@login_required()
def UnhideStock(request, pk):
    if not is_admin(request.user):
        messages.error(request, "You do not have permission to unhide inventory records.")
        return redirect('TotalStock')
    item = get_object_or_404(DaftarBarang, nomor=pk)
    item.status = 'Active'
    item.save()
    messages.success(request, f"'{item.nama_product}' has been made visible again.")
    return redirect('TotalStock')


@login_required()
def Cart(request):
    TransaksiListProdukFormset = formset_factory(TransaksiProductListForm, extra=1)
    formset = TransaksiListProdukFormset()
    data = DaftarBarang.objects.filter(user_id=request.user.profile.id, status='Active')

    if request.method == 'POST':
        formset_post = TransaksiListProdukFormset(request.POST)
        if formset_post.is_valid():
            total_harga_transaksi = 0
            total_jumlah_transaksi = 0
            transaksi = DaftarTransaksi.objects.create(user_id=request.user.profile.id)
            transaksi.save()
            for form in formset_post:
                if form.cleaned_data['quantity'] < 1:
                    messages.warning(request, 'Jumlah barang yang dibeli tidak boleh kosong!')
                    return redirect('/cart/')
                output = form.save(transaksi)
                if output is False:
                    messages.warning(request, 'Barang melebihi batas stock!')
                    return redirect('/cart/')
                total_harga_transaksi += output.quantity
                total_jumlah_transaksi += output.subtotal
            transaksi.total = total_jumlah_transaksi
            transaksi.produk_jumlah = total_harga_transaksi
            transaksi.save()
            messages.success(request, 'Transaksi Berhasil!')
            return HttpResponseRedirect('/struck/' + str(transaksi.nomor) + '/')
        else:
            messages.warning(request, 'Data yang dibeli tidak boleh kosong!')
            return redirect('/cart/')
    else:
        context = {
            'data_barang': data,
            'forms': formset,
            'request_user': request.user.profile.id,
        }
        return render(request, 'cashier/cart.html', context)


@login_required()
def StruckPembelian(request, pk):
    dataStruck = DaftarTransaksi.objects.get(nomor=pk)
    dataStruckListProduk = ListProductTransaksi.objects.filter(transaksi_id=dataStruck.nomor)
    dataUser = Profile.objects.get(id=dataStruck.user_id)

    import urllib.parse
    upi_qr_data = urllib.parse.quote(
        f"upi://pay?pa=merchant@upi&pn=JCB+Hydraulic+Parts&am={dataStruck.total}&cu=INR&tn=Invoice+%23{dataStruck.nomor}"
    )

    context = {
        'dataStruck': dataStruck,
        'dataStruckListProduk': dataStruckListProduk,
        'dataUser': dataUser,
        'upi_qr_data': upi_qr_data,
    }
    return render(request, 'cashier/struck.html', context)


@login_required()
def DaftarPembelian(request):
    data = DaftarTransaksi.objects.filter(user_id=request.user.profile.id)
    context = {
        'data': data,
    }
    return render(request, 'cashier/purchased.html', context)


@login_required()
def ReportView(request):
    # Handle saving Financial Insight metadata via POST
    if request.method == 'POST' and request.POST.get('action') == 'save_insight_meta':
        # Try to update the latest instead of creating a new one each time
        fi = FinancialInsightReport.objects.filter(user_id=request.user.profile.id).first()
        if not fi:
            fi = FinancialInsightReport(user_id=request.user.profile.id)
            
        fi.name = request.POST.get('fi_name', '')
        fi.mobile_number = request.POST.get('fi_mobile', '')
        fi_date = request.POST.get('fi_date', '')
        fi_time = request.POST.get('fi_time', '')
        
        try:
            fi.date = datetime.strptime(fi_date, '%Y-%m-%d').date() if fi_date else None
        except Exception:
            fi.date = None
        try:
            fi.time = datetime.strptime(fi_time, '%H:%M').time() if fi_time else None
        except Exception:
            fi.time = None
            
        fi.save()
        messages.success(request, 'Report Header Details updated successfully.')
        return redirect('ReportView')

    # Handle AJAX date-range filter
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('startDate'):
        endDate   = None
        startDate = request.GET.get('startDate')
        if not request.GET.get('endDate'):
            endDateConverter = datetime.today()
        else:
            endDate = request.GET.get('endDate')
            endDateConverter = datetime.strptime(endDate, '%Y-%m-%d').date()
        startDateConverter = datetime.strptime(startDate, '%Y-%m-%d').date()
        from_user    = DaftarTransaksi.objects.filter(
            user_id=request.user.profile.id,
            created__date__gte=startDateConverter,
            created__date__lte=endDateConverter,
        )
        daftar_barang = ListProductTransaksi.objects.filter(transaksi_id__in=from_user)
        # Get latest insight for the AJAX response too
        latest_insight = FinancialInsightReport.objects.filter(user_id=request.user.profile.id).first()
        return render(request, 'cashier/report_details.html', {
            'daftar_barang': daftar_barang, 
            'num': startDateConverter,
            'latest_insight': latest_insight
        })

    from_user     = DaftarTransaksi.objects.filter(user_id=request.user.profile.id)
    daftar_barang = ListProductTransaksi.objects.filter(transaksi_id__in=from_user)

    # Latest saved insight meta for this user
    latest_insight = FinancialInsightReport.objects.filter(user_id=request.user.profile.id).first()

    # Aggregated financial data for the insight panel
    inventory_qs     = DaftarBarang.objects.filter(user_id=request.user.profile.id, status='Active')
    total_stock_val  = sum(Decimal(d.subtotal_harga_beli or 0) for d in inventory_qs)
    total_sell_val   = sum(Decimal(d.subtotal_harga_jual or 0) for d in inventory_qs)
    potential_profit = total_sell_val - total_stock_val

    context = {
        'daftar_barang': daftar_barang,
        'from_user': from_user,
        'latest_insight': latest_insight,
        'total_stock_val': total_stock_val,
        'total_sell_val': total_sell_val,
        'potential_profit': potential_profit,
    }
    return render(request, 'cashier/report.html', context)


# ── Estimate Slip Views ────────────────────────────────────────────────────
@login_required()
def EstimateView(request):
    if not is_admin(request.user):
        messages.error(request, 'Only admin users can generate estimate slips.')
        return redirect('HomeIndex')

    # Last saved estimate to auto-number the next one
    last = EstimateSlip.objects.order_by('-id').first()
    next_no = f"EST-{(last.id + 1) if last else 1:04d}"

    saved_estimates = EstimateSlip.objects.filter(user_id=request.user.profile.id).order_by('-created_at')[:10]

    today = datetime.today()
    context = {
        'next_no': next_no,
        'today_date': today.strftime('%Y-%m-%d'),
        'today_time': today.strftime('%H:%M'),
        'saved_estimates': saved_estimates,
        'is_admin': True,
    }
    return render(request, 'cashier/estimate.html', context)


@login_required()
@require_GET
def PartPriceLookup(request):
    part_name = request.GET.get('name', '').strip()

    if not part_name:
        return JsonResponse({'price': '0.00'})

    # 1. Try to find in user's actual inventory first (DaftarBarang)
    inventory_item = DaftarBarang.objects.filter(
        user_id=request.user.profile.id,
        nama_product__iexact=part_name
    ).order_by('-created').first()

    if inventory_item:
        return JsonResponse({'price': str(inventory_item.harga_jual_satuan or '0.00')})

    # 2. Fallback to master catalog (JCBPart)
    part = JCBPart.objects.filter(
        Q(name__iexact=part_name) | Q(part_number__iexact=part_name)
    ).first()

    if part:
        return JsonResponse({'price': str(part.price)})

    return JsonResponse({'price': '0.00'})


@login_required()
@require_GET
def PartSuggestions(request):
    query = request.GET.get('q', '').strip()

    if not query:
        return JsonResponse({'results': []})

    # Search in Master Catalog
    catalog_parts = JCBPart.objects.filter(
        Q(name__icontains=query) | Q(part_number__icontains=query) | Q(description__icontains=query)
    ).order_by('name')[:10]

    # Search in User's actual Inventory
    inv_parts = DaftarBarang.objects.filter(
        user_id=request.user.profile.id,
        nama_product__icontains=query
    ).order_by('nama_product')[:10]

    results = []
    seen_names = set()

    # Prioritize Inventory items in results
    for p in inv_parts:
        if p.nama_product not in seen_names:
            results.append({
                'name': p.nama_product,
                'part_number': 'STOCK',
                'label': f"📦 [In Stock] {p.nama_product}",
            })
            seen_names.add(p.nama_product)

    for part in catalog_parts:
        if part.name not in seen_names:
            results.append({
                'name': part.name,
                'part_number': part.part_number,
                'label': f"{part.part_number} - {part.name}",
            })
            seen_names.add(part.name)

    return JsonResponse({'results': results[:15]})


@login_required()
@require_POST
def SaveEstimate(request):
    if not is_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON.'}, status=400)

    estimate_no   = data.get('estimate_no', '').strip()
    customer_name = data.get('customer_name', '').strip()
    mobile_number = data.get('mobile_number', '').strip()
    date_str      = data.get('date', '')
    time_str      = data.get('time', '')
    items         = data.get('items', [])

    if not customer_name or not estimate_no:
        return JsonResponse({'success': False, 'error': 'Estimate No. and Customer Name are required.'}, status=400)

    try:
        slip_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        slip_date = datetime.today().date()

    try:
        slip_time = datetime.strptime(time_str, '%H:%M').time()
    except Exception:
        slip_time = datetime.today().time()

    # Check for duplicate estimate_no and auto-suffix if needed
    if EstimateSlip.objects.filter(estimate_no=estimate_no).exists():
        estimate_no = f"{estimate_no}-{datetime.now().strftime('%S')}"

    slip = EstimateSlip.objects.create(
        user_id=request.user.profile.id,
        estimate_no=estimate_no,
        customer_name=customer_name,
        mobile_number=mobile_number,
        date=slip_date,
        time=slip_time,
    )

    total = Decimal('0')
    for row in items:
        qty  = Decimal(str(row.get('qty', 1)))
        rate = Decimal(str(row.get('rate', 0)))
        amt  = qty * rate
        total += amt
        EstimateSlipItem.objects.create(
            estimate=slip,
            particulars=row.get('particulars', ''),
            quantity=qty,
            rate=rate,
            amount=amt,
        )

    slip.total_amount = total
    slip.save()

    return JsonResponse({'success': True, 'id': slip.id, 'estimate_no': slip.estimate_no})
